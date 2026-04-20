from __future__ import annotations

from io import BytesIO
from pathlib import Path
import tempfile
import zipfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .compensation import build_waveform_diagnostic_exports
from .models import DatasetAnalysis, ExportArtifacts, ParsedMeasurement


def export_analysis_bundle(
    output_dir: str | Path,
    parsed_measurements: list[ParsedMeasurement],
    analyses: list[DatasetAnalysis],
    per_cycle_summary: pd.DataFrame,
    per_test_summary: pd.DataFrame,
    coverage: pd.DataFrame,
    config_snapshot_yaml: str,
    current_channel: str = "i_sum_signed",
    field_channel: str = "bz_mT",
    figures: dict[str, object] | None = None,
) -> ExportArtifacts:
    """Write all requested export artifacts into an output directory."""

    root_dir = Path(output_dir)
    root_dir.mkdir(parents=True, exist_ok=True)
    plots_dir = root_dir / "summary_plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    normalized_data_path = root_dir / "normalized_data.xlsx"
    per_test_summary_path = root_dir / "per_test_summary.xlsx"
    per_cycle_summary_path = root_dir / "per_cycle_summary.xlsx"
    report_path = root_dir / "analysis_report.md"
    config_snapshot_path = root_dir / "config_snapshot.yaml"
    excel_formatting_report_path = root_dir / "excel_formatting_check.md"

    mapping_table = _build_combined_mapping_table(parsed_measurements)
    raw_preview = _build_raw_preview(parsed_measurements)
    normalized_data = _build_normalized_data(parsed_measurements)
    waveform_fit_summary, representative_cycle_profiles = build_waveform_diagnostic_exports(
        analyses,
        current_channel=current_channel,
        field_channel=field_channel,
    )

    formatting_checks: dict[str, list[str]] = {}

    formatting_checks[normalized_data_path.name] = _write_excel_atomic(
        normalized_data_path,
        {
            "normalized_data": _excel_safe_frame(normalized_data),
            "raw_preview": _excel_safe_frame(raw_preview),
            "parsed_mapping": _excel_safe_frame(mapping_table),
            "representative_cycles": _excel_safe_frame(representative_cycle_profiles),
        },
    )

    formatting_checks[per_test_summary_path.name] = _write_excel_atomic(
        per_test_summary_path,
        {
            "per_test_summary": _excel_safe_frame(per_test_summary),
            "field_retention_table": _excel_safe_frame(_field_retention_table(per_test_summary)),
            "current_retention_table": _excel_safe_frame(_current_retention_table(per_test_summary)),
            "operating_map_table": _excel_safe_frame(_operating_map_table(per_test_summary)),
            "waveform_fit_summary": _excel_safe_frame(waveform_fit_summary),
            "coverage": _excel_safe_frame(coverage.reset_index()),
        },
    )

    formatting_checks[per_cycle_summary_path.name] = _write_excel_atomic(
        per_cycle_summary_path,
        {
            "per_cycle_summary": _excel_safe_frame(per_cycle_summary),
            "drift_table": _excel_safe_frame(_drift_table(per_cycle_summary)),
        },
    )

    config_snapshot_path.write_text(config_snapshot_yaml, encoding="utf-8")
    excel_formatting_report_path.write_text(
        _build_excel_formatting_report(formatting_checks),
        encoding="utf-8",
    )
    report_path.write_text(
        _build_analysis_report(
            parsed_measurements,
            analyses,
            per_cycle_summary,
            per_test_summary,
            waveform_fit_summary,
        ),
        encoding="utf-8",
    )

    if figures:
        for name, figure in figures.items():
            safe_name = name.replace(" ", "_").replace("/", "_")
            if hasattr(figure, "write_html"):
                figure.write_html(str(plots_dir / f"{safe_name}.html"))

    return ExportArtifacts(
        root_dir=root_dir,
        normalized_data_path=normalized_data_path,
        per_test_summary_path=per_test_summary_path,
        per_cycle_summary_path=per_cycle_summary_path,
        report_path=report_path,
        config_snapshot_path=config_snapshot_path,
        plots_dir=plots_dir,
        excel_formatting_report_path=excel_formatting_report_path,
    )


def build_export_zip_bytes(root_dir: str | Path) -> bytes:
    """Zip an export directory for Streamlit download."""

    root = Path(root_dir)
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in root.rglob("*"):
            if path.is_file():
                archive.write(path, arcname=path.relative_to(root))
    return buffer.getvalue()


def _build_combined_mapping_table(parsed_measurements: list[ParsedMeasurement]) -> pd.DataFrame:
    frames = []
    for parsed in parsed_measurements:
        rows = []
        for standard_field, source_column in parsed.mapping.items():
            rows.append(
                {
                    "source_file": parsed.source_file,
                    "sheet_name": parsed.sheet_name,
                    "standard_field": standard_field,
                    "source_column": source_column,
                }
            )
        frames.append(pd.DataFrame(rows))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _build_raw_preview(parsed_measurements: list[ParsedMeasurement]) -> pd.DataFrame:
    previews = []
    for parsed in parsed_measurements:
        preview = parsed.raw_frame.head(10).copy()
        preview.insert(0, "source_file", parsed.source_file)
        preview.insert(1, "sheet_name", parsed.sheet_name)
        previews.append(preview)
    return pd.concat(previews, ignore_index=True) if previews else pd.DataFrame()


def _build_normalized_data(parsed_measurements: list[ParsedMeasurement]) -> pd.DataFrame:
    frames = [parsed.normalized_frame for parsed in parsed_measurements]
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _field_retention_table(per_test_summary: pd.DataFrame) -> pd.DataFrame:
    columns = [
        column
        for column in per_test_summary.columns
        if column.startswith("achieved_b") and column.endswith("_pp_mean")
    ]
    base_columns = ["test_id", "waveform_type", "freq_hz", "current_pp_target_a"]
    return per_test_summary[base_columns + columns] if not per_test_summary.empty else pd.DataFrame()


def _current_retention_table(per_test_summary: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "test_id",
        "waveform_type",
        "freq_hz",
        "current_pp_target_a",
        "achieved_current_pp_a_mean",
        "current_retention",
        "reference_current_ratio",
    ]
    available = [column for column in columns if column in per_test_summary.columns]
    return per_test_summary[available] if not per_test_summary.empty else pd.DataFrame()


def _drift_table(per_cycle_summary: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "test_id",
        "cycle_index",
        "current_pp_drift_ratio",
        "temperature_drift_ratio",
        "gain_drift_ratio",
    ]
    field_drift_columns = [column for column in per_cycle_summary.columns if column.endswith("_pp_drift_ratio")]
    available = list(
        dict.fromkeys(
            column for column in columns + field_drift_columns if column in per_cycle_summary.columns
        )
    )
    return per_cycle_summary[available] if not per_cycle_summary.empty else pd.DataFrame()


def _operating_map_table(per_test_summary: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "test_id",
        "waveform_type",
        "freq_hz",
        "current_pp_target_a",
        "achieved_bz_mT_pp_mean",
        "achieved_bmag_mT_pp_mean",
        "temperature_rise_total_c",
        "warning_flags",
    ]
    available = [column for column in columns if column in per_test_summary.columns]
    return per_test_summary[available] if not per_test_summary.empty else pd.DataFrame()


def _build_analysis_report(
    parsed_measurements: list[ParsedMeasurement],
    analyses: list[DatasetAnalysis],
    per_cycle_summary: pd.DataFrame,
    per_test_summary: pd.DataFrame,
    waveform_fit_summary: pd.DataFrame,
) -> str:
    warning_count = sum(len(analysis.warnings) for analysis in analyses)
    lines = [
        "# 분석 보고서",
        "",
        f"- 입력 파일 수: {len(parsed_measurements)}",
        f"- 분석 테스트 수: {len(per_test_summary)}",
        f"- cycle 요약 행 수: {len(per_cycle_summary)}",
        f"- 경고 수: {warning_count}",
        "",
        "## 파일 목록",
    ]
    for parsed in parsed_measurements:
        lines.append(f"- `{parsed.source_file}` / `{parsed.sheet_name}` / `{parsed.normalized_frame['test_id'].iloc[0]}`")

    if not per_test_summary.empty:
        lines.extend(["", "## 테스트 요약", ""])
        for _, row in per_test_summary.iterrows():
            lines.append(
                "- "
                f"{row['test_id']}: freq={row['freq_hz']}, current_pp_target={row['current_pp_target_a']}, "
                f"field_pp_mean={row.get('achieved_bz_mT_pp_mean', row.get('achieved_bz_pp_mean', 'n/a'))}, "
                f"temp_rise={row.get('temperature_rise_total_c', 'n/a')}"
            )

    if not waveform_fit_summary.empty:
        lines.extend(["", "## waveform fit summary", ""])
        for _, row in waveform_fit_summary.iterrows():
            lines.append(
                "- "
                f"{row['test_id']}: "
                f"current_corr={row.get('current_shape_corr', 'n/a')}, "
                f"current_nrmse={row.get('current_shape_nrmse', 'n/a')}, "
                f"current_phase_lag_s={row.get('current_phase_lag_seconds', 'n/a')}, "
                f"voltage_corr={row.get('voltage_shape_corr', 'n/a')}"
            )

    if analyses:
        lines.extend(["", "## 경고", ""])
        for analysis in analyses:
            for warning in analysis.warnings:
                lines.append(f"- `{analysis.parsed.normalized_frame['test_id'].iloc[0]}`: {warning}")

    return "\n".join(lines) + "\n"


def _excel_safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    safe = frame.copy()
    for column in safe.columns:
        series = safe[column]
        if isinstance(series.dtype, pd.DatetimeTZDtype):
            safe[column] = series.dt.tz_localize(None)
            continue
        if pd.api.types.is_object_dtype(series):
            safe[column] = series.map(_excel_safe_value)
    return safe


def _excel_safe_value(value):
    if isinstance(value, pd.Timestamp) and value.tzinfo is not None:
        return value.tz_localize(None)
    return value


def _write_excel_atomic(path: Path, sheets: dict[str, pd.DataFrame]) -> list[str]:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_fd, temp_name = tempfile.mkstemp(
        dir=str(path.parent),
        prefix=f"{path.stem}_",
        suffix=path.suffix,
    )
    temp_path = Path(temp_name)
    import os

    os.close(temp_fd)
    try:
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            for sheet_name, frame in sheets.items():
                frame.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]
                _apply_excel_sheet_formatting(worksheet, frame, sheet_name)
        verification_lines = _verify_excel_workbook_formatting(temp_path, sheets)
        temp_path.replace(path)
        return verification_lines
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise


def _apply_excel_sheet_formatting(worksheet, frame: pd.DataFrame, sheet_name: str) -> None:
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.freeze_panes = _resolve_freeze_panes(sheet_name, frame)
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.zoomScale = 90
        worksheet.sheet_view.zoomScaleNormal = 90
        worksheet.sheet_view.showGridLines = True
    worksheet.sheet_view.tabSelected = False
    worksheet.row_dimensions[1].height = 22

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, column_name in enumerate(frame.columns, start=1):
        column_letter = get_column_letter(column_index)
        series = frame[column_name]
        worksheet.column_dimensions[column_letter].width = _estimate_excel_width(sheet_name, column_name, series)
        worksheet.column_dimensions[column_letter].bestFit = True
        number_format = _infer_excel_number_format(column_name, series)
        cell_alignment = _infer_excel_alignment(column_name, series)
        if number_format:
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.number_format = number_format
                cell.alignment = cell_alignment
        else:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).alignment = cell_alignment


def _estimate_excel_width(sheet_name: str, column_name: str, series: pd.Series) -> float:
    number_format = _infer_excel_number_format(column_name, series)
    min_width, max_width = _column_width_bounds(sheet_name, column_name, series)
    sample = series.head(250)
    text_lengths = [len(str(column_name))]
    text_lengths.extend(
        len(_render_excel_preview_value(value, number_format))
        for value in sample.tolist()
        if value not in (None, "")
    )
    estimated = max(text_lengths, default=min_width) + 2
    return float(min(max(estimated, min_width), max_width))


def _infer_excel_number_format(column_name: str, series: pd.Series) -> str | None:
    column_key = str(column_name).lower()
    if pd.api.types.is_datetime64_any_dtype(series):
        return "yyyy-mm-dd hh:mm:ss.000"
    if not pd.api.types.is_numeric_dtype(series):
        return None
    if any(token in column_key for token in ("index", "count", "expected", "sample")):
        return "0"
    if any(token in column_key for token in ("corr", "nrmse", "rmse", "phase")):
        return "0.0000"
    if any(token in column_key for token in ("freq", "time", "duration", "lag")):
        return "0.0000"
    if column_key.endswith("_ratio") or "retention" in column_key or column_key.startswith("reference_"):
        return "0.00%"
    if any(token in column_key for token in ("current", "voltage", "field", "temp", "gain", "amp", "bx", "by", "bz", "bmag")):
        return "0.000"
    return "0.000"


def _resolve_freeze_panes(sheet_name: str, frame: pd.DataFrame) -> str:
    if frame.shape[1] > 1:
        return "B2"
    return "A2"


def _column_width_bounds(sheet_name: str, column_name: str, series: pd.Series) -> tuple[int, int]:
    column_key = str(column_name).lower()
    sheet_max = {
        "normalized_data": 26,
        "raw_preview": 24,
        "parsed_mapping": 40,
        "representative_cycles": 20,
        "per_test_summary": 38,
        "field_retention_table": 24,
        "current_retention_table": 24,
        "operating_map_table": 26,
        "waveform_fit_summary": 24,
        "coverage": 18,
        "per_cycle_summary": 24,
        "drift_table": 22,
    }.get(sheet_name, 28)

    if pd.api.types.is_datetime64_any_dtype(series) or any(token in column_key for token in ("timestamp", "datetime", "date")):
        return 24, 30
    if any(token in column_key for token in ("warning", "note", "description", "log")):
        return 18, max(sheet_max, 56)
    if any(token in column_key for token in ("test_id", "source_file", "file_name", "path")):
        return 20, max(sheet_max, 42)
    if any(token in column_key for token in ("source_column", "standard_field", "mapping", "sheet_name")):
        return 14, max(sheet_max, 32)
    if any(token in column_key for token in ("waveform", "mode", "status", "axis", "channel")):
        return 12, max(sheet_max, 18)
    if any(token in column_key for token in ("ratio", "retention")) or column_key.startswith("reference_"):
        return 12, max(sheet_max, 16)
    if any(token in column_key for token in ("index", "count", "expected", "sample")):
        return 9, 12
    if any(token in column_key for token in ("freq", "time", "duration", "lag", "phase")):
        return 12, max(sheet_max, 20)
    if pd.api.types.is_numeric_dtype(series):
        return 12, max(sheet_max, 16)
    return 10, sheet_max


def _infer_excel_alignment(column_name: str, series: pd.Series) -> Alignment:
    column_key = str(column_name).lower()
    if any(token in column_key for token in ("warning", "note", "description", "log")):
        return Alignment(horizontal="left", vertical="top", wrap_text=True)
    if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
        return Alignment(horizontal="right", vertical="center")
    if any(token in column_key for token in ("waveform", "mode", "status", "sheet_name")):
        return Alignment(horizontal="center", vertical="center")
    return Alignment(horizontal="left", vertical="center")


def _render_excel_preview_value(value, number_format: str | None) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d %H:%M:%S.%f")[:23]
    if number_format == "0.00%":
        try:
            return f"{float(value) * 100:.2f}%"
        except (TypeError, ValueError):
            return str(value)
    if number_format == "0":
        try:
            return str(int(round(float(value))))
        except (TypeError, ValueError):
            return str(value)
    if number_format == "0.0000":
        try:
            return f"{float(value):.4f}"
        except (TypeError, ValueError):
            return str(value)
    if number_format == "0.000":
        try:
            return f"{float(value):.3f}"
        except (TypeError, ValueError):
            return str(value)
    return str(value)


def _verify_excel_workbook_formatting(path: Path, sheets: dict[str, pd.DataFrame]) -> list[str]:
    workbook = load_workbook(path)
    report_lines: list[str] = []
    failures: list[str] = []

    for sheet_name, frame in sheets.items():
        worksheet = workbook[sheet_name]
        expected_freeze = _resolve_freeze_panes(sheet_name, frame)
        actual_freeze = worksheet.freeze_panes or "None"
        filter_ref = worksheet.auto_filter.ref or "None"
        if actual_freeze != expected_freeze:
            failures.append(f"{path.name}/{sheet_name}: freeze_panes={actual_freeze} (expected {expected_freeze})")
        if worksheet.max_row >= 1 and not worksheet.auto_filter.ref:
            failures.append(f"{path.name}/{sheet_name}: auto_filter missing")

        width_samples: list[str] = []
        format_samples: list[str] = []
        for column_index, column_name in enumerate(frame.columns, start=1):
            column_letter = get_column_letter(column_index)
            expected_width = _estimate_excel_width(sheet_name, column_name, frame[column_name])
            actual_width = float(worksheet.column_dimensions[column_letter].width or 0)
            if abs(actual_width - expected_width) > 0.6:
                failures.append(
                    f"{path.name}/{sheet_name}:{column_name} width={actual_width:.1f} (expected {expected_width:.1f})"
                )
            if len(width_samples) < 4:
                width_samples.append(f"{column_name}={actual_width:.1f}")

            expected_format = _infer_excel_number_format(column_name, frame[column_name])
            if expected_format and worksheet.max_row >= 2:
                actual_format = worksheet.cell(row=2, column=column_index).number_format
                if actual_format != expected_format:
                    failures.append(
                        f"{path.name}/{sheet_name}:{column_name} number_format={actual_format} "
                        f"(expected {expected_format})"
                    )
                if len(format_samples) < 4:
                    format_samples.append(f"{column_name}={actual_format}")

        report_lines.append(
            "- "
            f"{path.name} / {sheet_name}: freeze={actual_freeze}, filter={filter_ref}, "
            f"widths=[{', '.join(width_samples)}], formats=[{', '.join(format_samples)}]"
        )

    if failures:
        raise ValueError("Excel formatting verification failed:\n" + "\n".join(failures))
    return report_lines


def _build_excel_formatting_report(formatting_checks: dict[str, list[str]]) -> str:
    lines = [
        "# Excel Formatting Check",
        "",
        "- Raw numeric values are preserved.",
        "- Display formatting, column widths, filters, and freeze panes are applied for readability.",
        "",
    ]
    for workbook_name, workbook_lines in formatting_checks.items():
        lines.append(f"## {workbook_name}")
        lines.extend(workbook_lines)
        lines.append("")
    return "\n".join(lines).strip() + "\n"
